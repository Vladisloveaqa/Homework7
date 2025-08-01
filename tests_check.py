import io
import zipfile
import csv
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import os
# Тест который проверяет в зипе файлы и их содержимое
def test_zip_contents():
    BASE_DIR = os.path.dirname(__file__)
    zip_path = os.path.join(BASE_DIR, "test.zip")

    with zipfile.ZipFile(zip_path, "r") as zf:

        # PDF
        with zf.open("pdfka.pdf") as pdf_file:
            pdf_reader = PdfReader(io.BytesIO(pdf_file.read()))
            assert len(pdf_reader.pages) == 1

        # XLSX
        with zf.open("summy.xlsx") as xlsx_file:
            wb = load_workbook(io.BytesIO(xlsx_file.read()))
            ws = wb.active
            assert ws["A1"].value == "Сумма руб."
            assert ws["B1"].value == "прпрп4544554"

        # CSV
        with zf.open("csv1.csv") as csv_file:
            text_stream = io.TextIOWrapper(csv_file, encoding="utf-8")
            reader = csv.reader(text_stream)
            rows = list(reader)
            assert rows[0] == ["Сумма руб."]

